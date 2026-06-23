#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
B站爬虫结果导出Excel工具
读取 bilibili_results 目录下的 JSON/CSV 结果文件
导出为简洁的 Excel，只保留 UID + 昵称
"""

import json, csv, glob, os, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 关键修复：双击运行时当前目录可能不是脚本所在目录
SCRIPT_DIR = Path(__file__).resolve().parent
RESULTS_DIR = SCRIPT_DIR / "bilibili_results"

def load_results():
    """从 JSON 和 CSV 文件中读取所有结果"""
    records = []
    seen = set()

    # 优先读 JSON（数据更完整）
    for jf in sorted(RESULTS_DIR.glob("result_*.json"), reverse=True):
        try:
            with open(jf, encoding="utf-8") as f:
                data = json.load(f)
            for r in data:
                uid = r.get("uid")
                if uid and uid not in seen:
                    seen.add(uid)
                    records.append({"uid": uid, "name": r.get("name", "")})
        except Exception:
            pass

    # 补充读 CSV（如果 JSON 漏了）
    for cf in sorted(RESULTS_DIR.glob("result_*.csv"), reverse=True):
        try:
            with open(cf, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    uid = int(r.get("uid", 0))
                    if uid and uid not in seen:
                        seen.add(uid)
                        records.append({"uid": uid, "name": r.get("name", "")})
        except Exception:
            pass

    return records

def export_excel(records, output_path):
    """导出为美观的 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "B站账号"

    # 样式定义
    header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    data_font = Font(name="微软雅黑", size=11)
    data_align = Alignment(horizontal="center", vertical="center")
    uid_align = Alignment(horizontal="center", vertical="center")
    name_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25

    # 表头
    ws.cell(row=1, column=1, value="UID").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=1).border = thin_border

    ws.cell(row=1, column=2, value="昵称").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = header_align
    ws.cell(row=1, column=2).border = thin_border

    # 数据行
    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for i, r in enumerate(records):
        row = i + 2
        uid_cell = ws.cell(row=row, column=1, value=r["uid"])
        uid_cell.font = data_font
        uid_cell.alignment = uid_align
        uid_cell.border = thin_border

        name_cell = ws.cell(row=row, column=2, value=r["name"])
        name_cell.font = data_font
        name_cell.alignment = name_align
        name_cell.border = thin_border

        # 奇偶行交替背景色
        if i % 2 == 1:
            uid_cell.fill = even_fill
            name_cell.fill = even_fill

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加筛选
    ws.auto_filter.ref = f"A1:B{len(records)+1}"

    wb.save(output_path)

def main():
    print("=" * 45)
    print("  B站爬虫结果 → Excel 导出工具")
    print("=" * 45)

    if not RESULTS_DIR.exists():
        print(f"\n未找到结果目录：{RESULTS_DIR}")
        print("请先运行爬虫生成结果文件")
        input("\n按回车退出...")
        return

    records = load_results()
    if not records:
        print("\n结果目录中没有找到数据")
        input("\n按回车退出...")
        return

    print(f"\n读取到 {len(records)} 条记录（已自动去重）")

    # 导出到结果目录
    output = RESULTS_DIR / "B站账号清单.xlsx"
    export_excel(records, output)

    print(f"\n导出完成！")
    print(f"  文件：{output}")
    print(f"  内容：{len(records)} 个账号的 UID + 昵称")
    print(f"  打开即可查看，表头冻结 + 奇偶行交替色")
    input("\n按回车退出...")

if __name__ == "__main__":
    main()
