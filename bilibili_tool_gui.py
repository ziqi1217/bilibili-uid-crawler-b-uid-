#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
B站UID爬虫 GUI版 v9
功能1：多线程爬取B站UID账号信息（等级0/关注0/粉丝0/非中文昵称）
功能2：导出爬取结果为Excel（仅UID+昵称）
使用 customtkinter 界面
"""

import re, time, random, csv, json, sys, threading, queue as qmod
from pathlib import Path
import requests as req_lib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import customtkinter as ctk
from tkinter import messagebox, filedialog

# ===== 路径定位（兼容PyInstaller打包） =====
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent

OUT_DIR = BASE_DIR / "bilibili_results"
PROGRESS_FILE = OUT_DIR / "_progress.json"

# ===== 爬虫配置 =====
API_URL    = "https://api.bilibili.com/x/space/acc/info"
TIMEOUT    = 10
DELAY_MIN  = 0.08
DELAY_MAX  = 0.20
AUTO_SAVE_EVERY = 15

UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Chrome/124.0.0.0 Safari/605.1",
]

# ===== 工具函数 =====
def has_chinese(text: str) -> bool:
    return bool(re.search(r'[\u4e00-\u9fff]', text))

def is_target(u: dict) -> bool:
    return (u.get('level', -1) == 0 and
            u.get('following', -1) == 0 and
            u.get('follower', -1) == 0 and
            not has_chinese(u.get('name', '')))

def fetch(uid: int, cookie_str: str):
    headers = {
        "User-Agent": random.choice(UA_POOL),
        "Referer": "https://www.bilibili.com/",
    }
    if cookie_str:
        headers["Cookie"] = cookie_str
    for attempt in range(3):
        try:
            r = req_lib.get(API_URL, params={"mid": uid}, headers=headers, timeout=TIMEOUT)
            d = r.json()
            if d.get("code") == 0:
                info = d["data"]
                return {
                    "uid": uid, "name": info.get("name", ""),
                    "level": info.get("level", 0),
                    "following": info.get("following", 0),
                    "follower": info.get("follower", 0),
                    "sign": info.get("sign", ""),
                    "jointime": info.get("jointime", 0),
                    "_status": "ok",
                }
            elif d.get("code") in (-352, -799):
                wait = random.uniform(3, 6)  # 限速后等待更久
                time.sleep(wait)
                continue
            else:
                return {"uid": uid, "_status": "skip"}
        except Exception as e:
            if attempt < 2:
                time.sleep(random.uniform(1.5, 3))
                continue
            return {"uid": uid, "_status": f"error:{type(e).__name__}"}
    return {"uid": uid, "_status": "rate_limited"}

def gen_uids(prefix: str, suffix: str, uid_len: int = 7):
    if not prefix and not suffix:
        raise ValueError("前缀和后缀至少填一个")
    if uid_len < 1 or uid_len > 10:
        raise ValueError("UID位数须在1~10之间")
    res = []
    if prefix and not suffix:
        pad = uid_len - len(prefix)
        if pad < 0:
            raise ValueError(f"前缀长度{len(prefix)}超过了UID位数{uid_len}")
        for i in range(10 ** pad):
            res.append(int(prefix + str(i).zfill(pad)))
    elif suffix and not prefix:
        pad = uid_len - len(suffix)
        if pad < 0:
            raise ValueError(f"后缀长度{len(suffix)}超过了UID位数{uid_len}")
        for i in range(10 ** pad):
            res.append(int(str(i).zfill(pad) + suffix))
    else:
        pre_pad = uid_len - len(prefix) - len(suffix)
        if pre_pad < 0:
            raise ValueError(f"前缀+后缀总长{len(prefix)+len(suffix)}超过了UID位数{uid_len}")
        for i in range(10 ** pre_pad):
            res.append(int(prefix + str(i).zfill(pre_pad) + suffix))
    return res

def format_time(sec):
    h, r = divmod(int(sec), 3600)
    m, s = divmod(r, 60)
    if h > 0:
        return f"{h}h{m}m{s}s"
    elif m > 0:
        return f"{m}m{s}s"
    else:
        return f"{s}s"

# ===== 保存 =====
def save_results_to_files(results_list, label=""):
    if not results_list:
        return
    OUT_DIR.mkdir(exist_ok=True)
    ts = int(time.time())
    csv_path  = OUT_DIR / f"result_{ts}{label}.csv"
    json_path = OUT_DIR / f"result_{ts}{label}.json"

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["uid","name","level","following","follower","sign","jointime"])
        for u in results_list:
            w.writerow([u["uid"],u["name"],u["level"],u["following"],u["follower"],u["sign"],u["jointime"]])

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results_list, f, ensure_ascii=False, indent=2)

    return csv_path

def save_progress(scanned_set, found_count):
    OUT_DIR.mkdir(exist_ok=True)
    with open(PROGRESS_FILE, "w") as f:
        json.dump({"scanned": sorted(scanned_set), "found_count": found_count}, f)

def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE) as f:
            d = json.load(f)
        return set(d.get("scanned", [])), d.get("found_count", 0)
    return set(), 0

# ===== Excel导出 =====
def load_results_from_files():
    records = []
    seen = set()
    for jf in sorted(OUT_DIR.glob("result_*.json"), reverse=True):
        try:
            with open(jf, encoding="utf-8") as f:
                data = json.load(f)
            for r in data:
                uid = r.get("uid")
                if uid and uid not in seen:
                    seen.add(uid)
                    records.append({"uid": uid, "name": r.get("name", "")})
        except Exception:
            pass
    for cf in sorted(OUT_DIR.glob("result_*.csv"), reverse=True):
        try:
            with open(cf, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    uid = int(r.get("uid", 0))
                    if uid and uid not in seen:
                        seen.add(uid)
                        records.append({"uid": uid, "name": r.get("name", "")})
        except Exception:
            pass
    return records

def do_export_excel(output_path):
    records = load_results_from_files()
    if not records:
        return 0

    wb = Workbook()
    ws = wb.active
    ws.title = "B站账号"

    header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="微软雅黑", size=11)
    uid_align = Alignment(horizontal="center", vertical="center")
    name_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25

    for col, val in enumerate(["UID", "昵称"], 1):
        c = ws.cell(row=1, column=col, value=val)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for i, r in enumerate(records):
        row = i + 2
        uid_cell = ws.cell(row=row, column=1, value=r["uid"])
        uid_cell.font = data_font
        uid_cell.alignment = uid_align
        uid_cell.border = thin_border
        name_cell = ws.cell(row=row, column=2, value=r["name"])
        name_cell.font = data_font
        name_cell.alignment = name_align
        name_cell.border = thin_border
        if i % 2 == 1:
            uid_cell.fill = even_fill
            name_cell.fill = even_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{len(records)+1}"
    wb.save(output_path)
    return len(records)


# ===== GUI =====
class BiliCrawlerApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # 窗口配置
        self.title("B站UID爬虫 v9")
        self.geometry("800x600")
        self.minsize(700, 500)
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        # 爬虫状态
        self.crawl_running = False
        self.crawl_results = []
        self.crawl_scanned = 0
        self.crawl_rate_limited = 0
        self.crawl_total = 0
        self.crawl_start_time = 0
        self.crawl_threads_list = []

        # 构建界面
        self._build_ui()

    def _build_ui(self):
        # ---- 顶部标题 ----
        title_frame = ctk.CTkFrame(self, fg_color="#00A1D6", height=50)
        title_frame.pack(fill="x", padx=0, pady=0)
        title_frame.pack_propagate(False)

        title_label = ctk.CTkLabel(title_frame, text="B站UID爬虫",
                                    font=ctk.CTkFont(size=22, weight="bold"),
                                    text_color="white")
        title_label.pack(side="left", padx=20, pady=10)

        subtitle = ctk.CTkLabel(title_frame, text="多线程极速版 | 等级0/关注0/粉丝0/非中文昵称",
                                 font=ctk.CTkFont(size=12),
                                 text_color="#E0E0E0")
        subtitle.pack(side="left", padx=10, pady=10)

        # ---- 左侧：参数设置 ----
        left_frame = ctk.CTkScrollableFrame(self, width=280, label_text="参数设置")
        left_frame.pack(side="left", fill="y", padx=(10,5), pady=10)
        # 不限制 propagate，允许滚动

        # UID位数（标题已由ScrollableFrame的label_text显示）
        self.uid_len_var = ctk.StringVar(value="7")
        self._add_param(left_frame, "UID位数", self.uid_len_var, "7/8/9/10")

        # 前缀
        self.prefix_var = ctk.StringVar()
        self._add_param(left_frame, "UID前缀", self.prefix_var, "可不填")

        # 后缀
        self.suffix_var = ctk.StringVar()
        self._add_param(left_frame, "UID后缀", self.suffix_var, "可不填")

        # 线程数
        self.threads_var = ctk.StringVar(value="8")
        self._add_param(left_frame, "线程数", self.threads_var, "1~30，推荐8~15")

        # Cookie
        self.cookie_var = ctk.StringVar()
        cookie_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        cookie_frame.pack(fill="x", padx=15, pady=(5,2))
        ctk.CTkLabel(cookie_frame, text="Cookie SESSDATA",
                      font=ctk.CTkFont(size=13)).pack(anchor="w")
        ctk.CTkLabel(cookie_frame, text="不填则慢，填了提速5~10倍",
                      font=ctk.CTkFont(size=10), text_color="#888").pack(anchor="w")

        self.cookie_entry = ctk.CTkEntry(left_frame, textvariable=self.cookie_var,
                                          placeholder_text="粘贴SESSDATA值...",
                                          height=35, width=240)
        self.cookie_entry.pack(padx=15, pady=(2,10))

        # 按钮
        btn_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        btn_frame.pack(fill="x", padx=15, pady=10)

        self.start_btn = ctk.CTkButton(btn_frame, text="▶ 开始爬取",
                                         fg_color="#00A1D6", hover_color="#0088B5",
                                         font=ctk.CTkFont(size=14, weight="bold"),
                                         height=40, command=self._on_start)
        self.start_btn.pack(fill="x", pady=(0,5))

        self.stop_btn = ctk.CTkButton(btn_frame, text="■ 停止",
                                       fg_color="#E74C3C", hover_color="#C0392B",
                                       font=ctk.CTkFont(size=14, weight="bold"),
                                       height=40, command=self._on_stop,
                                       state="disabled")
        self.stop_btn.pack(fill="x", pady=(0,5))

        self.export_btn = ctk.CTkButton(btn_frame, text="↓ 导出Excel",
                                         fg_color="#27AE60", hover_color="#1E8449",
                                         font=ctk.CTkFont(size=14, weight="bold"),
                                         height=40, command=self._on_export)
        self.export_btn.pack(fill="x", pady=(0,5))

        # ---- 右侧：进度+结果 ----
        right_frame = ctk.CTkFrame(self)
        right_frame.pack(side="right", fill="both", expand=True, padx=(5,10), pady=10)

        # 进度区
        progress_frame = ctk.CTkFrame(right_frame)
        progress_frame.pack(fill="x", padx=10, pady=(10,5))

        self.progress_label = ctk.CTkLabel(progress_frame, text="就绪",
                                             font=ctk.CTkFont(size=14, weight="bold"))
        self.progress_label.pack(pady=(10,5), padx=10, anchor="w")

        self.progress_bar = ctk.CTkProgressBar(progress_frame, height=15)
        self.progress_bar.pack(fill="x", padx=10, pady=(0,5))
        self.progress_bar.set(0)

        # 统计区
        stats_frame = ctk.CTkFrame(progress_frame, fg_color="transparent")
        stats_frame.pack(fill="x", padx=10, pady=(0,10))

        self.stat_scanned = ctk.CTkLabel(stats_frame, text="已扫描: 0",
                                           font=ctk.CTkFont(size=12))
        self.stat_scanned.pack(side="left", padx=(0,15))

        self.stat_found = ctk.CTkLabel(stats_frame, text="已找到: 0",
                                        font=ctk.CTkFont(size=12), text_color="#27AE60")
        self.stat_found.pack(side="left", padx=(0,15))

        self.stat_speed = ctk.CTkLabel(stats_frame, text="速度: 0/s",
                                        font=ctk.CTkFont(size=12), text_color="#00A1D6")
        self.stat_speed.pack(side="left", padx=(0,15))

        self.stat_time = ctk.CTkLabel(stats_frame, text="剩余: --",
                                       font=ctk.CTkFont(size=12))
        self.stat_time.pack(side="left")

        self.stat_limited = ctk.CTkLabel(stats_frame, text="限速: 0",
                                          font=ctk.CTkFont(size=12), text_color="#E74C3C")
        self.stat_limited.pack(side="left", padx=(15,0))

        # Cookie状态标签
        self.cookie_status = ctk.CTkLabel(progress_frame, text="⚠ 无Cookie（速度慢）",
                                            font=ctk.CTkFont(size=11), text_color="#E67E22")
        self.cookie_status.pack(anchor="w", padx=10, pady=(0,5))

        # 结果表格区
        table_label = ctk.CTkLabel(right_frame, text="找到的账号",
                                    font=ctk.CTkFont(size=14, weight="bold"))
        table_label.pack(pady=(5,2), padx=10, anchor="w")

        # 滚动文本框作为结果展示
        self.result_text = ctk.CTkTextbox(right_frame, height=300,
                                           font=ctk.CTkFont(size=12))
        self.result_text.pack(fill="both", expand=True, padx=10, pady=(2,10))
        self.result_text.insert("1.0", "等待开始...\n")
        self.result_text.configure(state="disabled")

    def _add_param(self, parent, label_text, var, hint):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.pack(fill="x", padx=15, pady=(5,2))
        ctk.CTkLabel(frame, text=label_text, font=ctk.CTkFont(size=13)).pack(anchor="w")
        ctk.CTkLabel(frame, text=hint, font=ctk.CTkFont(size=10), text_color="#888").pack(anchor="w")
        entry = ctk.CTkEntry(parent, textvariable=var, height=35, width=240)
        entry.pack(padx=15, pady=(2,5))

    def _log(self, msg):
        self.result_text.configure(state="normal")
        self.result_text.insert("end", msg + "\n")
        self.result_text.configure(state="disabled")
        # 自动滚动到底部
        self.result_text.yview("end")

    def _on_start(self):
        if self.crawl_running:
            return

        # 获取参数
        prefix = self.prefix_var.get().strip()
        suffix = self.suffix_var.get().strip()
        cookie_raw = self.cookie_var.get().strip()

        if cookie_raw and "=" not in cookie_raw:
            cookie_str = f"SESSDATA={cookie_raw}"
        else:
            cookie_str = cookie_raw

        uid_len = 7
        try:
            uid_len = int(self.uid_len_var.get().strip())
        except ValueError:
            uid_len = 7
        uid_len = max(1, min(10, uid_len))

        n_threads = 8
        try:
            n_threads = int(self.threads_var.get().strip())
        except ValueError:
            n_threads = 8
        n_threads = max(1, min(30, n_threads))

        # 更新Cookie状态
        if cookie_str:
            self.cookie_status.configure(text="🚀 Cookie模式（速度快）", text_color="#27AE60")
        else:
            self.cookie_status.configure(text="⚠ 无Cookie（速度慢）", text_color="#E67E22")

        # 生成UID列表
        if not prefix and not suffix:
            messagebox.showwarning("参数错误", "前缀和后缀至少填一个！")
            return

        try:
            uid_list = gen_uids(prefix, suffix, uid_len)
        except ValueError as e:
            messagebox.showerror("参数错误", str(e))
            return

        if len(uid_list) > 500000:
            if not messagebox.askyesno("确认", f"将扫描 {len(uid_list):,} 个UID，可能耗时较长。\n确认继续？"):
                return

        # 断点续传
        scanned_set, prev_found = load_progress()
        if scanned_set:
            if messagebox.askyesno("断点续传", f"发现上次未完成任务（已扫描{len(scanned_set)}个，找到{prev_found}个）\n跳过已扫描继续？"):
                uid_list = [u for u in uid_list if u not in scanned_set]
            else:
                scanned_set.clear()

        # 重置状态
        self.crawl_running = True
        self.crawl_results = []
        self.crawl_scanned = 0
        self.crawl_rate_limited = 0
        self.crawl_total = len(uid_list)
        self.crawl_start_time = time.time()

        # UI切换
        self.start_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.progress_bar.set(0)
        self.progress_label.configure(text=f"扫描中... {len(uid_list):,} 个UID")
        self.result_text.configure(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.insert("1.0", f"开始扫描 {len(uid_list):,} 个UID | {uid_len}位数 | {n_threads}线程 | Cookie={'ON' if cookie_str else 'OFF'}\n\n")
        self.result_text.configure(state="disabled")

        OUT_DIR.mkdir(exist_ok=True)

        # 启动爬虫线程
        task_q = qmod.Queue()
        for u in uid_list:
            task_q.put(u)

        def worker_func(tid):
            rate_count = 0
            while self.crawl_running:
                try:
                    uid = task_q.get(timeout=1)
                except qmod.Empty:
                    break

                user = fetch(uid, cookie_str)

                # 处理错误/限速状态
                status = (user or {}).get("_status", "ok")
                if status == "rate_limited":
                    rate_count += 1
                    self.crawl_rate_limited += 1
                    if rate_count <= 3:
                        self._log(f"  ⚠ UID={uid} 被限速，等待重试...")
                    elif rate_count == 5:
                        self._log(f"  ⚠ 持续限速中，建议降低线程数或稍后再试")
                    self.crawl_scanned += 1
                    task_q.task_done()
                    time.sleep(random.uniform(2, 4))
                    continue
                elif status.startswith("error"):
                    self._log(f"  ❌ UID={uid} 网络错误: {status}")
                    self.crawl_scanned += 1
                    task_q.task_done()
                    continue

                rate_count = 0  # 重置计数
                self.crawl_scanned += 1

                if user and is_target(user):
                    self.crawl_results.append(user)
                    n = len(self.crawl_results)
                    self._log(f"  ✅ #{n}: UID={user['uid']}  昵称={user['name']}")

                task_q.task_done()
                time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

        self.crawl_threads_list = []
        for i in range(n_threads):
            t = threading.Thread(target=worker_func, args=(i+1), daemon=True)
            t.start()
            self.crawl_threads_list.append(t)

        # 启动进度更新线程
        self._update_progress()

    def _update_progress(self):
        if not self.crawl_running:
            return

        s = self.crawl_scanned
        total = self.crawl_total
        f_count = len(self.crawl_results)
        elapsed = time.time() - self.crawl_start_time
        speed = s / elapsed if elapsed > 0 else 0
        pct = s / total * 100 if total > 0 else 0
        remain = (total - s) / speed if speed > 0 else 0

        # 更新UI
        self.progress_bar.set(pct / 100)
        self.stat_scanned.configure(text=f"已扫描: {s:,}")
        self.stat_found.configure(text=f"已找到: {f_count}")
        self.stat_speed.configure(text=f"速度: {speed:.1f}/s")
        self.stat_time.configure(text=f"剩余: {format_time(remain)}")
        self.stat_limited.configure(text=f"限速: {self.crawl_rate_limited}")

        # 自动保存
        if f_count >= AUTO_SAVE_EVERY and f_count % AUTO_SAVE_EVERY == 0:
            save_results_to_files(self.crawl_results)
            save_progress(set(), f_count)

        # 检查是否完成
        if s >= total or not any(t.is_alive() for t in self.crawl_threads_list):
            self._on_crawl_done()
            return

        # 继续更新
        self.after(500, self._update_progress)

    def _on_crawl_done(self):
        # 等线程结束
        for t in self.crawl_threads_list:
            t.join(timeout=3)

        # 保存结果
        save_results_to_files(self.crawl_results)
        save_progress(set(), len(self.crawl_results))

        elapsed = time.time() - self.crawl_start_time
        speed = self.crawl_scanned / elapsed if elapsed > 0 else 0

        # 更新UI
        self.crawl_running = False
        self.start_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")
        self.progress_bar.set(1.0)

        self.progress_label.configure(
            text=f"完成！用时 {format_time(elapsed)} | 找到 {len(self.crawl_results)} 个账号")

        self._log(f"\n{'='*45}")
        self._log(f"扫描完成！")
        self._log(f"  用时：{format_time(elapsed)}")
        self._log(f"  扫描：{self.crawl_scanned:,} 个UID")
        self._log(f"  找到：{len(self.crawl_results)} 个目标账号")
        self._log(f"  平均速度：{speed:.1f} uid/s")
        self._log(f"  结果保存在：{OUT_DIR}")
        self._log(f"{'='*45}")

    def _on_stop(self):
        if not self.crawl_running:
            return
        self.crawl_running = False

        # 保存已找到的结果
        if self.crawl_results:
            save_results_to_files(self.crawl_results, "_中断保存")

        elapsed = time.time() - self.crawl_start_time

        self.start_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")

        self.progress_label.configure(text=f"已停止 | 找到 {len(self.crawl_results)} 个账号")
        self._log(f"\n手动停止，已保存 {len(self.crawl_results)} 条结果")

    def _on_export(self):
        if not OUT_DIR.exists():
            messagebox.showwarning("无数据", "未找到结果目录，请先运行爬虫！")
            return

        records = load_results_from_files()
        if not records:
            messagebox.showwarning("无数据", "结果目录中没有数据！")
            return

        output = OUT_DIR / "B站账号清单.xlsx"
        count = do_export_excel(output)

        if count > 0:
            messagebox.showinfo("导出成功", f"已导出 {count} 个账号到\n{output}")
            self._log(f"\n导出Excel完成：{count} 个账号 -> {output}")
        else:
            messagebox.showwarning("导出失败", "没有数据可导出")


if __name__ == "__main__":
    app = BiliCrawlerApp()
    app.mainloop()
