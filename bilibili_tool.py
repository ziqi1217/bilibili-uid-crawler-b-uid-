#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
B站UID爬虫 整合版 v8
功能1：多线程爬取B站UID账号信息（等级0/关注0/粉丝0/非中文昵称）
功能2：导出爬取结果为Excel（仅UID+昵称）
双击运行后弹出菜单选择功能
"""

import re, time, random, csv, json, sys, signal, threading, queue as qmod
from pathlib import Path
import requests as req_lib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ===== 路径定位（兼容PyInstaller打包） =====
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent

OUT_DIR = BASE_DIR / "bilibili_results"
PROGRESS_FILE = OUT_DIR / "_progress.json"

# ===== 爬虫配置 =====
API_URL    = "https://api.bilibili.com/x/space/acc/info"
TIMEOUT    = 10
DELAY_MIN  = 0.08
DELAY_MAX  = 0.20
THREADS    = 8
AUTO_SAVE_EVERY = 15

UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Chrome/124.0.0.0 Safari/605.1",
]

# ===== 全局状态 =====
results     = []
found_lock  = threading.Lock()
scanned     = 0
scan_lock   = threading.Lock()
total_uids  = 0
running     = True
cookie_str  = ""
start_time  = 0

# ===== 工具函数 =====
def has_chinese(text: str) -> bool:
    return bool(re.search(r'[\u4e00-\u9fff]', text))

def is_target(u: dict) -> bool:
    return (u.get('level', -1) == 0 and
            u.get('following', -1) == 0 and
            u.get('follower', -1) == 0 and
            not has_chinese(u.get('name', '')))

def fetch(uid: int):
    headers = {
        "User-Agent": random.choice(UA_POOL),
        "Referer": "https://www.bilibili.com/",
    }
    if cookie_str:
        headers["Cookie"] = cookie_str
    for attempt in range(3):
        try:
            r = req_lib.get(API_URL, params={"mid": uid}, headers=headers, timeout=TIMEOUT)
            d = r.json()
            if d.get("code") == 0:
                info = d["data"]
                return {
                    "uid": uid, "name": info.get("name", ""),
                    "level": info.get("level", 0),
                    "following": info.get("following", 0),
                    "follower": info.get("follower", 0),
                    "sign": info.get("sign", ""),
                    "jointime": info.get("jointime", 0),
                }
            elif d.get("code") in (-352, -799):
                time.sleep(random.uniform(2, 4))
                continue
            else:
                return None
        except Exception:
            if attempt < 2:
                time.sleep(random.uniform(1, 2))
                continue
            return None
    return None

def gen_uids(prefix: str, suffix: str, uid_len: int = 7):
    if not prefix and not suffix:
        raise ValueError("前缀和后缀至少填一个")
    if uid_len < 1 or uid_len > 10:
        raise ValueError("UID位数须在1~10之间")
    res = []
    if prefix and not suffix:
        pad = uid_len - len(prefix)
        if pad < 0:
            raise ValueError(f"前缀长度{len(prefix)}超过了UID位数{uid_len}")
        for i in range(10 ** pad):
            res.append(int(prefix + str(i).zfill(pad)))
    elif suffix and not prefix:
        pad = uid_len - len(suffix)
        if pad < 0:
            raise ValueError(f"后缀长度{len(suffix)}超过了UID位数{uid_len}")
        for i in range(10 ** pad):
            res.append(int(str(i).zfill(pad) + suffix))
    else:
        pre_pad = uid_len - len(prefix) - len(suffix)
        if pre_pad < 0:
            raise ValueError(f"前缀+后缀总长{len(prefix)+len(suffix)}超过了UID位数{uid_len}")
        for i in range(10 ** pre_pad):
            res.append(int(prefix + str(i).zfill(pre_pad) + suffix))
    return res

def format_time(sec):
    h, r = divmod(int(sec), 3600)
    m, s = divmod(r, 60)
    if h > 0:
        return f"{h}时{m}分{s}秒"
    elif m > 0:
        return f"{m}分{s}秒"
    else:
        return f"{s}秒"

# ===== 保存 =====
def save_results(label=""):
    with found_lock:
        local = list(results)
    if not local:
        return
    ts = int(time.time())
    csv_path  = OUT_DIR / f"result_{ts}{label}.csv"
    json_path = OUT_DIR / f"result_{ts}{label}.json"

    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["uid","name","level","following","follower","sign","jointime"])
        for u in local:
            w.writerow([u["uid"],u["name"],u["level"],u["following"],u["follower"],u["sign"],u["jointime"]])

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(local, f, ensure_ascii=False, indent=2)

    print(f"\n  已保存 {len(local)} 条 -> {csv_path.name}")

def save_progress(scanned_set):
    OUT_DIR.mkdir(exist_ok=True)
    with open(PROGRESS_FILE, "w") as f:
        json.dump({"scanned": sorted(scanned_set), "found_count": len(results)}, f)

def load_progress():
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE) as f:
            d = json.load(f)
        return set(d.get("scanned", [])), d.get("found_count", 0)
    return set(), 0

# ===== 线程工作函数 =====
def worker(uid_queue: qmod.Queue, thread_id: int):
    global scanned
    while running:
        try:
            uid = uid_queue.get(timeout=1)
        except qmod.Empty:
            break

        user = fetch(uid)

        with scan_lock:
            scanned += 1

        if user and is_target(user):
            with found_lock:
                results.append(user)
                n = len(results)
            print(f"  [T{thread_id}] #{n}: UID={user['uid']} 昵称={user['name']}")

        uid_queue.task_done()
        time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

# ===== Ctrl+C =====
def sigint_handler(signum, frame):
    global running
    running = False
    print("\n\n收到中断，正在保存结果...")
    save_results("_中断保存")
    print("保存完成。程序退出。")
    sys.exit(0)

signal.signal(signal.SIGINT, sigint_handler)

# ===== 功能1：爬虫 =====
def run_crawler():
    global total_uids, running, cookie_str, start_time, results, scanned

    print("=" * 55)
    print("  B站UID爬虫 - 多线程极速版")
    print(f"  默认 {THREADS} 线程并发 | Ctrl+C 安全退出")
    print("=" * 55)

    uid_len_input = input("\nUID位数（7/8/9/10，回车=7）: ").strip()
    uid_len = 7
    if uid_len_input.isdigit():
        uid_len = max(1, min(10, int(uid_len_input)))

    prefix = input("UID前缀（可不填）: ").strip()
    suffix = input("UID后缀（可不填）: ").strip()
    cookie = input("B站Cookie SESSDATA值（不填则慢）: ").strip()

    if cookie and "=" not in cookie:
        cookie = f"SESSDATA={cookie}"
    cookie_str = cookie

    t_input = input(f"线程数（回车=默认{THREADS}）: ").strip()
    n_threads = THREADS
    if t_input.isdigit():
        n_threads = max(1, min(30, int(t_input)))

    try:
        uid_list = gen_uids(prefix, suffix, uid_len)
    except ValueError as e:
        print(f"错误：{e}")
        return

    if len(uid_list) > 500000:
        print(f"\n⚠ 警告：将扫描 {len(uid_list):,} 个UID，可能耗时较长")
        r = input("确认继续？(y/n): ").strip().lower()
        if r != "y":
            return

    # 断点续传
    scanned_set, prev_found = load_progress()
    if scanned_set:
        print(f"\n发现上次未完成任务（已扫描 {len(scanned_set)} 个，找到 {prev_found} 个）")
        r = input("跳过已扫描继续？(y/n): ").strip().lower()
        if r == "y":
            uid_list = [u for u in uid_list if u not in scanned_set]
            print(f"  剩余 {len(uid_list):,} 个UID")
        else:
            scanned_set.clear()

    total_uids = len(uid_list)
    OUT_DIR.mkdir(exist_ok=True)

    print(f"\n{'='*55}")
    print(f"开始扫描 {total_uids:,} 个UID | {uid_len}位数 | {n_threads}线程 | Cookie={'ON' if cookie_str else 'OFF'}")
    print(f"提示：按 Ctrl+C 随时中断并自动保存")
    print(f"{'='*55}")

    start_time = time.time()
    last_t = start_time
    last_save = 0

    task_q = qmod.Queue()
    for u in uid_list:
        task_q.put(u)

    threads = []
    for i in range(n_threads):
        t = threading.Thread(target=worker, args=(task_q, i+1), daemon=True)
        t.start()
        threads.append(t)

    while True:
        with scan_lock:
            s = scanned

        if s >= total_uids or not any(t.is_alive() for t in threads):
            break

        now = time.time()
        if now - last_t >= 1.5:
            elapsed = now - start_time
            speed = s / elapsed if elapsed > 0 else 0
            pct = s / total_uids * 100
            remain = (total_uids - s) / speed if speed > 0 else 0

            bar_len = 40
            fill = int(pct / 100 * bar_len)
            bar = "█" * fill + "░" * (bar_len - fill)

            with found_lock:
                f_count = len(results)

            print(f"  [{bar}] {pct:.1f}% | {s:,}/{total_uids:,} | "
                  f"{speed:.1f}/s | 找到{f_count} | 剩余{format_time(remain)}")
            last_t = now

            if f_count >= AUTO_SAVE_EVERY and f_count - last_save >= AUTO_SAVE_EVERY:
                save_results()
                save_progress(set())
                last_save = f_count

        time.sleep(0.3)

    for t in threads:
        t.join(timeout=5)

    save_results()
    save_progress(set())

    elapsed = time.time() - start_time
    avg_spd = scanned / elapsed if elapsed > 0 else 0

    print(f"\n{'='*55}")
    print(f"扫描完成！")
    print(f"  用时：{format_time(elapsed)}")
    print(f"  扫描：{scanned:,} 个UID")
    print(f"  找到：{len(results)} 个目标账号")
    print(f"  平均速度：{avg_spd:.1f} uid/s")
    print(f"  结果保存在：{OUT_DIR}")
    print(f"{'='*55}")

    # 询问是否导出Excel
    r = input("\n是否导出Excel？(y/n): ").strip().lower()
    if r == "y":
        run_export()

# ===== 功能2：导出Excel =====
def load_results_from_files():
    """从 JSON 和 CSV 文件中读取所有结果"""
    records = []
    seen = set()

    for jf in sorted(OUT_DIR.glob("result_*.json"), reverse=True):
        try:
            with open(jf, encoding="utf-8") as f:
                data = json.load(f)
            for r in data:
                uid = r.get("uid")
                if uid and uid not in seen:
                    seen.add(uid)
                    records.append({"uid": uid, "name": r.get("name", "")})
        except Exception:
            pass

    for cf in sorted(OUT_DIR.glob("result_*.csv"), reverse=True):
        try:
            with open(cf, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for r in reader:
                    uid = int(r.get("uid", 0))
                    if uid and uid not in seen:
                        seen.add(uid)
                        records.append({"uid": uid, "name": r.get("name", "")})
        except Exception:
            pass

    return records

def export_excel(records, output_path):
    """导出为美观的 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "B站账号"

    header_font = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    data_font = Font(name="微软雅黑", size=11)
    uid_align = Alignment(horizontal="center", vertical="center")
    name_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25

    for col, val in enumerate(["UID", "昵称"], 1):
        c = ws.cell(row=1, column=col, value=val)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for i, r in enumerate(records):
        row = i + 2
        uid_cell = ws.cell(row=row, column=1, value=r["uid"])
        uid_cell.font = data_font
        uid_cell.alignment = uid_align
        uid_cell.border = thin_border

        name_cell = ws.cell(row=row, column=2, value=r["name"])
        name_cell.font = data_font
        name_cell.alignment = name_align
        name_cell.border = thin_border

        if i % 2 == 1:
            uid_cell.fill = even_fill
            name_cell.fill = even_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{len(records)+1}"

    wb.save(output_path)

def run_export():
    print("\n" + "=" * 45)
    print("  B站爬虫结果 -> Excel 导出")
    print("=" * 45)

    if not OUT_DIR.exists():
        print(f"\n未找到结果目录：{OUT_DIR}")
        print("请先运行爬虫生成结果文件")
        return

    records = load_results_from_files()
    if not records:
        print("\n结果目录中没有找到数据")
        return

    print(f"\n读取到 {len(records)} 条记录（已自动去重）")

    output = OUT_DIR / "B站账号清单.xlsx"
    export_excel(records, output)

    print(f"\n导出完成！")
    print(f"  文件：{output}")
    print(f"  内容：{len(records)} 个账号的 UID + 昵称")

# ===== 主菜单 =====
def main():
    while True:
        print("\n" + "=" * 55)
        print("  B站UID爬虫 整合版 v8")
        print("=" * 55)
        print("\n  1. 开始爬取B站UID")
        print("  2. 导出结果为Excel")
        print("  0. 退出")
        print("\n" + "-" * 55)

        choice = input("请选择功能 (0/1/2): ").strip()

        if choice == "1":
            run_crawler()
        elif choice == "2":
            run_export()
        elif choice == "0":
            print("\n再见！")
            break
        else:
            print("\n无效选择，请输入 0、1 或 2")

if __name__ == "__main__":
    main()
